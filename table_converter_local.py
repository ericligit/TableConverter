"""
Table Image → Spreadsheet Converter  (Local / No API Key)
Uses img2table + EasyOCR to extract tables entirely on-device.
Exports to CSV, Excel (.xlsx), or Google Sheets.
"""

import os
import csv
import json
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    HAS_DND = True
except ImportError:
    HAS_DND = False

try:
    from PIL import Image, ImageTk, ImageEnhance, ImageFilter
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import gspread
    from google.oauth2.service_account import Credentials
    HAS_GSPREAD = True
except ImportError:
    HAS_GSPREAD = False


# ─────────────────────────────────────────────
#  Image preprocessing
# ─────────────────────────────────────────────

def to_png(image_path: str) -> str:
    """Convert image to a temporary PNG without any other processing."""
    import tempfile
    img = Image.open(image_path).convert("RGB")
    tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    img.save(tmp.name, "PNG")
    tmp.close()
    return tmp.name


def to_grayscale_png(image_path: str) -> str:
    """
    Convert image to high-contrast grayscale.

    Colored table backgrounds (e.g. alternating red/green/yellow rows) create
    false vertical edges that mislead img2table's border detector, producing
    phantom columns.  Stripping color removes those false edges so that only
    the real grid lines remain visible.
    """
    import tempfile
    img = Image.open(image_path).convert("L")          # true grayscale
    img = ImageEnhance.Contrast(img).enhance(2.0)      # make borders pop
    tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    img.save(tmp.name, "PNG")
    tmp.close()
    return tmp.name


def preprocess_image(image_path: str) -> str:
    """
    Upscale 2x, sharpen, and boost contrast so EasyOCR reads small
    text more reliably. Returns path to a temporary PNG.

    NOTE: only use this when the original resolution is too small for OCR.
    The 2x upscale can break img2table's fixed-kernel border detection,
    so always try the original image first.
    """
    import tempfile
    img = Image.open(image_path).convert("RGB")

    # Upscale 2x with high-quality resampling
    img = img.resize((img.width * 2, img.height * 2), Image.LANCZOS)

    # Unsharp mask to crisp up glyph edges
    img = img.filter(ImageFilter.UnsharpMask(radius=1.5, percent=180, threshold=2))

    # Boost contrast slightly
    img = ImageEnhance.Contrast(img).enhance(1.4)

    # Boost sharpness
    img = ImageEnhance.Sharpness(img).enhance(2.0)

    tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    img.save(tmp.name, "PNG")
    tmp.close()
    return tmp.name


# ─────────────────────────────────────────────
#  Table extraction (local OCR)
# ─────────────────────────────────────────────

def _easyocr_direct(image_path: str) -> list[list[str]]:
    """
    Reconstruct a table purely from EasyOCR text bounding-box positions —
    no border / line detection involved.

    Works for any table regardless of cell background colour (heat-map
    tables, alternating-row colour tables, etc.) because colour never
    enters the algorithm.

    Algorithm
    ---------
    1. Run EasyOCR → (bbox, text, confidence) triples.
    2. Cluster text blocks into rows by vertical (y-centre) proximity,
       using average glyph height as the tolerance.
    3. Find column anchor positions from the densest row (usually the header).
    4. Assign every text block to its nearest column anchor.
    """
    import easyocr

    reader = easyocr.Reader(["en"], gpu=False, verbose=False)
    results = reader.readtext(image_path, detail=1, paragraph=False)

    if not results:
        return []

    items = []
    for bbox, text, conf in results:
        text = text.strip()
        if not text or conf < 0.15:
            continue
        xs = [p[0] for p in bbox]
        ys = [p[1] for p in bbox]
        items.append({
            "text": text,
            "cx":   (min(xs) + max(xs)) / 2,
            "cy":   (min(ys) + max(ys)) / 2,
            "h":    max(ys) - min(ys),
        })

    if not items:
        return []

    # ── Cluster into rows ────────────────────────────────────────────────────
    items.sort(key=lambda i: i["cy"])
    avg_h   = sum(i["h"] for i in items) / len(items)
    row_tol = avg_h * 0.65      # items within this Δy share a row

    rows_raw: list[list[dict]] = [[items[0]]]
    for item in items[1:]:
        row_cy = sum(i["cy"] for i in rows_raw[-1]) / len(rows_raw[-1])
        if abs(item["cy"] - row_cy) <= row_tol:
            rows_raw[-1].append(item)
        else:
            rows_raw.append([item])

    for row in rows_raw:
        row.sort(key=lambda i: i["cx"])

    # ── Determine column anchor positions from the fullest row ───────────────
    anchor   = max(rows_raw, key=len)
    col_cx   = [item["cx"] for item in anchor]
    n_cols   = len(col_cx)

    if n_cols == 0:
        return []

    def _nearest(cx: float) -> int:
        return min(range(n_cols), key=lambda j: abs(col_cx[j] - cx))

    # ── Build grid ───────────────────────────────────────────────────────────
    table: list[list[str]] = []
    for row in rows_raw:
        cells = [""] * n_cols
        for item in row:
            j = _nearest(item["cx"])
            cells[j] = (cells[j] + " " + item["text"]).strip() if cells[j] else item["text"]
        table.append(cells)

    return table


def extract_tables_local(image_path: str,
                         progress_cb=None) -> list[list[list[str]]]:
    """
    Returns a list of tables; each table is a list-of-lists of strings.

    Strategy order
    --------------
    1. EasyOCR-direct  — position-based clustering; immune to cell colours.
    2. img2table + grayscale PNG — grayscale removes colour noise from the
       border detector; good for tables with visible grid lines.
    3. img2table + borderless mode — for tables without visible borders.
    4. img2table + upscaled image — last resort for very small images.

    EasyOCR downloads its model (~100 MB) on first run.
    """
    import os
    from img2table.document import Image as Img2Image
    from img2table.ocr import EasyOCR as Img2EasyOCR

    # ── Strategy 1: EasyOCR-direct ───────────────────────────────────────────
    # Reconstructs table structure from text positions alone; no border
    # detection → unaffected by coloured cell backgrounds.
    if progress_cb:
        progress_cb("Detecting table via text-position clustering (pass 1/4)…")
    try:
        table = _easyocr_direct(image_path)
        if table:
            return [table]
    except Exception:
        pass

    # ── img2table fallback strategies ────────────────────────────────────────
    if progress_cb:
        progress_cb("Loading img2table OCR model (pass 2/4)…")
    ocr = Img2EasyOCR(lang=["en"])

    def _run(proc_path, borderless):
        doc = Img2Image(src=proc_path)
        return doc.extract_tables(
            ocr=ocr,
            implicit_rows=True,
            implicit_columns=True,
            borderless_tables=borderless,
            min_confidence=30,
        )

    def _parse(result):
        tables = []
        for tbl in (result if isinstance(result, list) else result.values()):
            if tbl is None:
                continue
            df = tbl.df.fillna("")
            rows = [list(df.columns.astype(str))] + df.values.tolist()
            rows = [[str(c) for c in row] for row in rows]
            if rows and all(h.isdigit() for h in rows[0]):
                rows = rows[1:]
            tables.append(rows)
        return tables

    # Strategy 2: grayscale — strips colour noise so only real grid lines remain
    if progress_cb:
        progress_cb("Retrying with grayscale image (pass 2/4)…")
    proc2 = to_grayscale_png(image_path)
    try:
        tables = _parse(_run(proc2, borderless=False))
    finally:
        os.unlink(proc2)
    if tables:
        return tables

    # Strategy 3: borderless mode — for tables without visible borders
    if progress_cb:
        progress_cb("Retrying in borderless mode (pass 3/4)…")
    proc3 = to_png(image_path)
    try:
        tables = _parse(_run(proc3, borderless=True))
    finally:
        os.unlink(proc3)
    if tables:
        return tables

    # Strategy 4: upscaled + sharpened — for very small / low-res source images
    if progress_cb:
        progress_cb("Retrying with upscaled image (pass 4/4)…")
    proc4 = preprocess_image(image_path)
    try:
        tables = _parse(_run(proc4, borderless=False))
    finally:
        os.unlink(proc4)

    return tables


# ─────────────────────────────────────────────
#  Export helpers
# ─────────────────────────────────────────────

def save_csv(rows: list[list[str]], out_path: str):
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)


def save_excel(rows: list[list[str]], out_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Table Data"

    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_fill = PatternFill("solid", fgColor="4472C4")
    h_font = Font(bold=True, color="FFFFFF")

    for r_i, row in enumerate(rows, 1):
        for c_i, val in enumerate(row, 1):
            cell = ws.cell(row=r_i, column=c_i, value=val)
            cell.border    = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if r_i <= 2:
                cell.fill = h_fill
                cell.font = h_font

    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(width + 4, 40)

    wb.save(out_path)


def upload_google_sheets(rows: list[list[str]], title: str,
                         creds_path: str) -> str:
    scopes = ["https://spreadsheets.google.com/feeds",
              "https://www.googleapis.com/auth/drive"]
    creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
    gc    = gspread.authorize(creds)
    sh    = gc.create(title)
    sh.get_worksheet(0).update(rows)
    sh.share(None, perm_type="anyone", role="reader")
    return sh.url


# ─────────────────────────────────────────────
#  GUI
# ─────────────────────────────────────────────

COLORS = dict(
    bg="#1e1e2e", surface="#313244", overlay="#45475a",
    text="#cdd6f4", subtext="#a6adc8", accent="#89b4fa",
    green="#a6e3a1", red="#f38ba8", yellow="#f9e2af",
    teal="#94e2d5",
)


class App(TkinterDnD.Tk if HAS_DND else tk.Tk):

    def __init__(self):
        super().__init__()
        self.title("Table Image → Spreadsheet  (Local OCR)")
        self.geometry("980x760")
        self.minsize(720, 580)
        self.configure(bg=COLORS["bg"])

        self.image_path  = tk.StringVar()
        self.export_fmt  = tk.StringVar(value="csv")
        self.gs_creds    = tk.StringVar()
        self.gs_title    = tk.StringVar(value="Extracted Table")
        self.status_msg  = tk.StringVar(value="Drop an image here or click Browse")
        self.table_var   = tk.IntVar(value=0)

        self.all_tables: list[list[list[str]]] = []   # all detected tables
        self.rows:        list[list[str]]       = []   # currently selected

        self._apply_style()
        self._build_ui()

    # ── Style ────────────────────────────────

    def _apply_style(self):
        s = ttk.Style(self)
        s.theme_use("clam")
        bg, fg, acc = COLORS["bg"], COLORS["text"], COLORS["accent"]
        surf = COLORS["surface"]

        s.configure("TFrame",       background=bg)
        s.configure("TLabel",       background=bg, foreground=fg,
                    font=("Segoe UI", 10))
        s.configure("Header.TLabel", background=bg, foreground=acc,
                    font=("Segoe UI", 13, "bold"))
        s.configure("Status.TLabel", background=bg, foreground=COLORS["yellow"],
                    font=("Segoe UI", 9))
        s.configure("TButton",      background=acc, foreground=bg,
                    font=("Segoe UI", 10, "bold"), padding=7, relief="flat")
        s.map("TButton",            background=[("active", COLORS["teal"]),
                                                ("disabled", COLORS["overlay"])])
        s.configure("TEntry",       fieldbackground=surf, foreground=fg,
                    insertcolor=fg, relief="flat", borderwidth=1)
        s.configure("TRadiobutton", background=bg, foreground=fg,
                    font=("Segoe UI", 10))
        s.configure("Treeview",     background=surf, foreground=fg,
                    fieldbackground=surf, rowheight=26, font=("Segoe UI", 9))
        s.configure("Treeview.Heading", background=COLORS["overlay"],
                    foreground=fg, font=("Segoe UI", 9, "bold"), relief="flat")
        s.map("Treeview",           background=[("selected", COLORS["overlay"])])
        s.configure("Horizontal.TProgressbar",
                    troughcolor=surf, background=acc, thickness=6)
        s.configure("Vertical.TScrollbar",
                    background=surf, troughcolor=bg, arrowcolor=fg)
        s.configure("Horizontal.TScrollbar",
                    background=surf, troughcolor=bg, arrowcolor=fg)

    # ── UI construction ───────────────────────

    def _build_ui(self):
        p = dict(padx=16, pady=6)

        # Title
        ttk.Label(self, text="Table Image → Spreadsheet",
                  style="Header.TLabel").pack(anchor="w", padx=16, pady=(12, 2))
        ttk.Label(self, text="Fully local — no API key required",
                  foreground=COLORS["green"]).pack(anchor="w", padx=16)

        ttk.Separator(self, orient="horizontal").pack(fill="x", padx=16, pady=6)

        # ── Drop zone ─────────────────────────
        self.drop_outer = tk.Frame(self, bg=COLORS["surface"],
                                   highlightbackground=COLORS["accent"],
                                   highlightthickness=2)
        self.drop_outer.pack(fill="x", padx=16, pady=4)

        self.preview_lbl = tk.Label(
            self.drop_outer,
            text="⬇  Drop image here   or   click Browse",
            bg=COLORS["surface"], fg=COLORS["accent"],
            font=("Segoe UI", 14), pady=44, cursor="hand2",
        )
        self.preview_lbl.pack(fill="both", expand=True)

        if HAS_DND:
            self.drop_outer.drop_target_register(DND_FILES)
            self.drop_outer.dnd_bind("<<Drop>>", self._on_drop)
        self.drop_outer.bind("<Button-1>",  lambda _: self._browse())
        self.preview_lbl.bind("<Button-1>", lambda _: self._browse())

        # ── Browse row ────────────────────────
        row = ttk.Frame(self)
        row.pack(fill="x", **p)
        ttk.Button(row, text="Browse…", command=self._browse).pack(side="left")
        ttk.Label(row, textvariable=self.image_path,
                  foreground=COLORS["green"]).pack(side="left", padx=8)

        # ── Export options ────────────────────
        exp = ttk.Frame(self)
        exp.pack(fill="x", **p)
        ttk.Label(exp, text="Export as:").pack(side="left")
        for val, lbl in [("csv", "CSV"), ("excel", "Excel (.xlsx)"),
                         ("gsheet", "Google Sheets")]:
            ttk.Radiobutton(exp, text=lbl, variable=self.export_fmt,
                            value=val, command=self._toggle_gs
                            ).pack(side="left", padx=10)

        # ── Google Sheets fields (hidden by default) ──
        self.gs_frame = ttk.Frame(self)
        ttk.Label(self.gs_frame, text="Sheet title:").grid(
            row=0, column=0, sticky="w")
        ttk.Entry(self.gs_frame, textvariable=self.gs_title, width=32
                  ).grid(row=0, column=1, padx=8, pady=2)
        ttk.Label(self.gs_frame, text="Service-account JSON:").grid(
            row=1, column=0, sticky="w")
        ttk.Entry(self.gs_frame, textvariable=self.gs_creds, width=42
                  ).grid(row=1, column=1, padx=8, pady=2)
        ttk.Button(self.gs_frame, text="Browse…",
                   command=self._browse_creds).grid(row=1, column=2, padx=4)

        # ── Action row ────────────────────────
        act = ttk.Frame(self)
        act.pack(fill="x", **p)
        self.extract_btn = ttk.Button(act, text="⚡  Extract Table",
                                      command=self._start_extraction)
        self.extract_btn.pack(side="left")
        self.progress = ttk.Progressbar(act, mode="indeterminate",
                                        length=220, style="Horizontal.TProgressbar")
        self.progress.pack(side="left", padx=12)

        # ── Table selector (shown when >1 table found) ──
        self.sel_frame = ttk.Frame(self)
        ttk.Label(self.sel_frame, text="Multiple tables found — select:").pack(
            side="left")
        self.sel_combo = ttk.Combobox(self.sel_frame, state="readonly", width=32)
        self.sel_combo.pack(side="left", padx=8)
        self.sel_combo.bind("<<ComboboxSelected>>", self._on_table_select)

        # ── Status ────────────────────────────
        ttk.Label(self, textvariable=self.status_msg,
                  style="Status.TLabel").pack(anchor="w", padx=16)

        # ── Data preview ─────────────────────
        pf = ttk.Frame(self)
        pf.pack(fill="both", expand=True, padx=16, pady=(4, 4))
        ttk.Label(pf, text="Extracted Data Preview:",
                  font=("Segoe UI", 10, "bold")).pack(anchor="w")

        tree_wrap = ttk.Frame(pf)
        tree_wrap.pack(fill="both", expand=True)
        self.tree = ttk.Treeview(tree_wrap, show="headings")
        vsb = ttk.Scrollbar(tree_wrap, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_wrap, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right",  fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)

        # ── Save / Upload ─────────────────────
        self.save_btn = ttk.Button(self, text="💾  Save / Upload",
                                   command=self._save, state="disabled")
        self.save_btn.pack(pady=(4, 14))

    # ── Event handlers ────────────────────────

    def _toggle_gs(self):
        if self.export_fmt.get() == "gsheet":
            self.gs_frame.pack(fill="x", padx=16, pady=2,
                               before=self.extract_btn.master)
        else:
            self.gs_frame.pack_forget()

    def _browse(self):
        path = filedialog.askopenfilename(
            title="Select table image",
            filetypes=[("Images", "*.jpg *.jpeg *.png *.gif *.webp *.bmp *.tiff"),
                       ("All files", "*.*")])
        if path:
            self._load_image(path)

    def _browse_creds(self):
        path = filedialog.askopenfilename(
            filetypes=[("JSON", "*.json"), ("All files", "*.*")])
        if path:
            self.gs_creds.set(path)

    def _on_drop(self, event):
        path = event.data.strip().strip("{}")
        if path:
            self._load_image(path)

    def _load_image(self, path: str):
        self.image_path.set(path)
        self.status_msg.set(f"Loaded: {Path(path).name}")
        self.all_tables = []
        self.rows = []
        self._clear_tree()
        self.sel_frame.pack_forget()
        self.save_btn.config(state="disabled")

        if not HAS_PIL:
            return
        try:
            img = Image.open(path)
            img.thumbnail((920, 230), Image.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            self.preview_lbl.config(image=photo, text="", pady=0)
            self.preview_lbl.image = photo
        except Exception:
            pass

    def _start_extraction(self):
        path = self.image_path.get()
        if not path or not Path(path).exists():
            messagebox.showwarning("No image", "Please select an image file first.")
            return
        self.extract_btn.config(state="disabled")
        self.save_btn.config(state="disabled")
        self._clear_tree()
        self.sel_frame.pack_forget()
        self.progress.start(10)
        self.status_msg.set("Starting local OCR…")
        threading.Thread(target=self._do_extraction,
                         args=(path,), daemon=True).start()

    def _do_extraction(self, path: str):
        try:
            tables = extract_tables_local(
                path, progress_cb=lambda m: self.after(0, self.status_msg.set, m)
            )
            self.after(0, self._extraction_done, tables)
        except Exception as exc:
            self.after(0, self._extraction_error, str(exc))

    def _extraction_done(self, tables: list[list[list[str]]]):
        self.progress.stop()
        self.extract_btn.config(state="normal")
        self.all_tables = tables

        if not tables:
            self.status_msg.set(
                "No tables detected. Try a cleaner image or one with visible borders.")
            messagebox.showwarning(
                "No tables found",
                "img2table could not detect a table in this image.\n\n"
                "Tips:\n• Ensure the table has clear border lines\n"
                "• Try cropping to just the table area\n"
                "• Higher-resolution images work better")
            return

        self.rows = tables[0]
        self._populate_tree(self.rows)
        self.save_btn.config(state="normal")

        if len(tables) > 1:
            options = [f"Table {i+1}  ({len(t)} rows × {max(len(r) for r in t)} cols)"
                       for i, t in enumerate(tables)]
            self.sel_combo["values"] = options
            self.sel_combo.current(0)
            self.sel_frame.pack(fill="x", padx=16, pady=2,
                                before=self.status_msg.master)
            self.status_msg.set(
                f"{len(tables)} tables detected. Showing Table 1.")
        else:
            t = tables[0]
            self.status_msg.set(
                f"Extracted {len(t)} rows × {max(len(r) for r in t)} columns. "
                "Ready to save.")

    def _extraction_error(self, msg: str):
        self.progress.stop()
        self.extract_btn.config(state="normal")
        self.status_msg.set(f"Error: {msg}")
        messagebox.showerror("Extraction failed",
                             f"{msg}\n\nMake sure img2table and easyocr are installed:\n"
                             "  pip install img2table easyocr")

    def _on_table_select(self, _event=None):
        idx = self.sel_combo.current()
        self.rows = self.all_tables[idx]
        self._populate_tree(self.rows)
        t = self.rows
        self.status_msg.set(
            f"Table {idx+1}: {len(t)} rows × {max(len(r) for r in t)} cols.")

    def _clear_tree(self):
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = []

    def _populate_tree(self, rows: list[list[str]]):
        self._clear_tree()
        if not rows:
            return
        n_cols = max(len(r) for r in rows)
        cols   = [str(i) for i in range(n_cols)]
        self.tree["columns"] = cols
        headers = rows[0] if rows else [""] * n_cols
        for c, h in zip(cols, headers):
            self.tree.heading(c, text=h or f"Col {c}")
            self.tree.column(c, width=max(80, len(str(h)) * 9 + 20),
                             anchor="center", minwidth=60)
        for row in rows[1:]:
            padded = list(row) + [""] * (n_cols - len(row))
            self.tree.insert("", "end", values=padded)

    # ── Save / upload ─────────────────────────

    def _save(self):
        if not self.rows:
            return
        {"csv": self._save_csv,
         "excel": self._save_excel,
         "gsheet": self._upload_gsheet}[self.export_fmt.get()]()

    def _save_csv(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".csv", filetypes=[("CSV", "*.csv")],
            initialfile="table.csv")
        if not path:
            return
        try:
            save_csv(self.rows, path)
            self.status_msg.set(f"Saved → {path}")
            messagebox.showinfo("Saved", f"CSV saved:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _save_excel(self):
        if not HAS_OPENPYXL:
            messagebox.showerror("Missing package", "pip install openpyxl")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile="table.xlsx")
        if not path:
            return
        try:
            save_excel(self.rows, path)
            self.status_msg.set(f"Saved → {path}")
            messagebox.showinfo("Saved", f"Excel saved:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _upload_gsheet(self):
        if not HAS_GSPREAD:
            messagebox.showerror("Missing package",
                                 "pip install gspread google-auth")
            return
        creds = self.gs_creds.get().strip()
        if not creds or not Path(creds).exists():
            messagebox.showwarning(
                "Credentials",
                "Provide a service-account JSON file.\n"
                "See: console.cloud.google.com → IAM → Service Accounts")
            return
        title = self.gs_title.get().strip() or "Extracted Table"
        self.status_msg.set("Uploading to Google Sheets…")
        self.save_btn.config(state="disabled")
        threading.Thread(target=self._do_upload,
                         args=(title, creds), daemon=True).start()

    def _do_upload(self, title: str, creds: str):
        try:
            url = upload_google_sheets(self.rows, title, creds)
            self.after(0, self._upload_done, url)
        except Exception as e:
            self.after(0, self._upload_error, str(e))

    def _upload_done(self, url: str):
        self.save_btn.config(state="normal")
        self.status_msg.set(f"Uploaded → {url}")
        messagebox.showinfo("Google Sheets", f"Created:\n{url}")

    def _upload_error(self, msg: str):
        self.save_btn.config(state="normal")
        self.status_msg.set(f"Upload error: {msg}")
        messagebox.showerror("Upload failed", msg)


# ─────────────────────────────────────────────
if __name__ == "__main__":
    app = App()
    app.mainloop()
