"""
Table Image to Spreadsheet Converter
Converts table images to Google Sheets, Excel, or CSV using Claude Vision API
"""

import os
import sys
import csv
import json
import base64
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    HAS_DND = True
except ImportError:
    HAS_DND = False

try:
    from PIL import Image, ImageTk
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

try:
    import anthropic
    HAS_ANTHROPIC = True
except ImportError:
    HAS_ANTHROPIC = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import gspread
    from google.oauth2.service_account import Credentials
    HAS_GSPREAD = True
except ImportError:
    HAS_GSPREAD = False


# ─────────────────────────────────────────────
#  Claude Vision extraction
# ─────────────────────────────────────────────

def extract_table_from_image(image_path: str, api_key: str) -> list[list[str]]:
    """Send image to Claude and extract table as list-of-lists."""
    client = anthropic.Anthropic(api_key=api_key)

    with open(image_path, "rb") as f:
        image_data = base64.standard_b64encode(f.read()).decode("utf-8")

    suffix = Path(image_path).suffix.lower()
    media_map = {".jpg": "image/jpeg", ".jpeg": "image/jpeg",
                 ".png": "image/png", ".gif": "image/gif",
                 ".webp": "image/webp"}
    media_type = media_map.get(suffix, "image/jpeg")

    prompt = (
        "This image contains a table. Extract ALL data from the table exactly as shown. "
        "Return ONLY a JSON array of arrays — no markdown, no explanation. "
        "Each inner array is one row (header rows first, then data rows). "
        "Preserve merged-cell headers by repeating the header value across the columns it spans. "
        "Empty cells should be empty strings. "
        "Example format: [[\"Header1\",\"Header2\"],[\"val1\",\"val2\"]]"
    )

    message = client.messages.create(
        model="claude-opus-4-6",
        max_tokens=4096,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": media_type,
                            "data": image_data,
                        },
                    },
                    {"type": "text", "text": prompt},
                ],
            }
        ],
    )

    raw = message.content[0].text.strip()
    # Strip markdown code fences if present
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    raw = raw.strip()
    return json.loads(raw)


# ─────────────────────────────────────────────
#  Export helpers
# ─────────────────────────────────────────────

def save_csv(rows: list[list[str]], out_path: str):
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def save_excel(rows: list[list[str]], out_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Table Data"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, cell_val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=cell_val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if r_idx <= 2:  # treat first two rows as headers
                cell.fill = header_fill
                cell.font = header_font

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(out_path)


def upload_google_sheets(rows: list[list[str]], sheet_title: str,
                         creds_path: str) -> str:
    scopes = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
    gc = gspread.authorize(creds)
    sh = gc.create(sheet_title)
    ws = sh.get_worksheet(0)
    ws.update(rows)
    sh.share(None, perm_type="anyone", role="reader")
    return sh.url


# ─────────────────────────────────────────────
#  GUI
# ─────────────────────────────────────────────

class App(TkinterDnD.Tk if HAS_DND else tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Table Image → Spreadsheet")
        self.geometry("900x720")
        self.minsize(700, 560)
        self.configure(bg="#1e1e2e")

        self.image_path = tk.StringVar()
        self.api_key    = tk.StringVar(value=os.environ.get("ANTHROPIC_API_KEY", ""))
        self.export_fmt = tk.StringVar(value="csv")
        self.gs_creds   = tk.StringVar()
        self.gs_title   = tk.StringVar(value="Extracted Table")
        self.status_msg = tk.StringVar(value="Drop an image or click Browse")
        self.rows: list[list[str]] = []

        self._build_ui()
        self._style()

    # ── UI construction ──────────────────────

    def _style(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        bg, fg, accent = "#1e1e2e", "#cdd6f4", "#89b4fa"
        style.configure("TFrame",       background=bg)
        style.configure("TLabel",       background=bg, foreground=fg, font=("Segoe UI", 10))
        style.configure("TButton",      background=accent, foreground="#1e1e2e",
                        font=("Segoe UI", 10, "bold"), padding=6)
        style.map("TButton",            background=[("active", "#74c7ec")])
        style.configure("TEntry",       fieldbackground="#313244", foreground=fg,
                        insertcolor=fg, borderwidth=0)
        style.configure("TRadiobutton", background=bg, foreground=fg,
                        font=("Segoe UI", 10))
        style.configure("Treeview",     background="#313244", foreground=fg,
                        fieldbackground="#313244", rowheight=24)
        style.configure("Treeview.Heading", background="#45475a", foreground=fg,
                        font=("Segoe UI", 10, "bold"))
        style.map("Treeview",           background=[("selected", "#585b70")])
        style.configure("TNotebook",    background=bg)
        style.configure("TNotebook.Tab", background="#313244", foreground=fg,
                        font=("Segoe UI", 10), padding=[10, 4])
        style.map("TNotebook.Tab",      background=[("selected", accent)],
                  foreground=[("selected", "#1e1e2e")])

    def _build_ui(self):
        root_pad = {"padx": 16, "pady": 8}

        # ── Top: API key ─────────────────────
        top = ttk.Frame(self)
        top.pack(fill="x", **root_pad)
        ttk.Label(top, text="Anthropic API Key:").pack(side="left")
        key_entry = ttk.Entry(top, textvariable=self.api_key, width=52, show="*")
        key_entry.pack(side="left", padx=(8, 0), fill="x", expand=True)
        ttk.Button(top, text="👁", width=3,
                   command=lambda: key_entry.config(
                       show="" if key_entry.cget("show") else "*")
                   ).pack(side="left", padx=(4, 0))

        # ── Drop zone ────────────────────────
        self.drop_frame = tk.Frame(self, bg="#313244", bd=2, relief="groove",
                                   cursor="hand2")
        self.drop_frame.pack(fill="x", padx=16, pady=4)

        self.preview_label = tk.Label(
            self.drop_frame,
            text="⬇  Drop image here  or  click Browse",
            bg="#313244", fg="#89b4fa",
            font=("Segoe UI", 14), pady=40
        )
        self.preview_label.pack(fill="both", expand=True)

        if HAS_DND:
            self.drop_frame.drop_target_register(DND_FILES)
            self.drop_frame.dnd_bind("<<Drop>>", self._on_drop)
        self.drop_frame.bind("<Button-1>", lambda _: self._browse())
        self.preview_label.bind("<Button-1>", lambda _: self._browse())

        # ── Browse + path ─────────────────────
        mid = ttk.Frame(self)
        mid.pack(fill="x", **root_pad)
        ttk.Button(mid, text="Browse…", command=self._browse).pack(side="left")
        ttk.Label(mid, textvariable=self.image_path,
                  foreground="#a6e3a1").pack(side="left", padx=8)

        # ── Export options ────────────────────
        opts = ttk.Frame(self)
        opts.pack(fill="x", **root_pad)
        ttk.Label(opts, text="Export as:").pack(side="left")
        for val, label in [("csv", "CSV"), ("excel", "Excel (.xlsx)"),
                           ("gsheet", "Google Sheets")]:
            ttk.Radiobutton(opts, text=label, variable=self.export_fmt,
                            value=val, command=self._toggle_gs
                            ).pack(side="left", padx=8)

        # ── Google Sheets extra fields ────────
        self.gs_frame = ttk.Frame(self)
        ttk.Label(self.gs_frame, text="Spreadsheet title:").grid(row=0, column=0, sticky="w")
        ttk.Entry(self.gs_frame, textvariable=self.gs_title, width=30
                  ).grid(row=0, column=1, padx=8, pady=2)
        ttk.Label(self.gs_frame, text="Service-account JSON:").grid(row=1, column=0, sticky="w")
        ttk.Entry(self.gs_frame, textvariable=self.gs_creds, width=40
                  ).grid(row=1, column=1, padx=8, pady=2)
        ttk.Button(self.gs_frame, text="Browse…",
                   command=self._browse_creds).grid(row=1, column=2, padx=4)

        # ── Convert button ────────────────────
        btn_row = ttk.Frame(self)
        btn_row.pack(fill="x", **root_pad)
        self.convert_btn = ttk.Button(btn_row, text="⚡  Extract & Convert",
                                      command=self._start_conversion)
        self.convert_btn.pack(side="left")
        self.progress = ttk.Progressbar(btn_row, mode="indeterminate", length=200)
        self.progress.pack(side="left", padx=12)

        # ── Status ────────────────────────────
        ttk.Label(self, textvariable=self.status_msg,
                  foreground="#f38ba8").pack(anchor="w", padx=16)

        # ── Preview table ─────────────────────
        preview_frame = ttk.Frame(self)
        preview_frame.pack(fill="both", expand=True, padx=16, pady=(4, 12))
        ttk.Label(preview_frame, text="Extracted Data Preview:",
                  font=("Segoe UI", 10, "bold")).pack(anchor="w")

        tree_outer = ttk.Frame(preview_frame)
        tree_outer.pack(fill="both", expand=True)

        self.tree = ttk.Treeview(tree_outer, show="headings")
        vsb = ttk.Scrollbar(tree_outer, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_outer, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right",  fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)

        # ── Save button ───────────────────────
        self.save_btn = ttk.Button(self, text="💾  Save / Upload",
                                   command=self._save, state="disabled")
        self.save_btn.pack(pady=(0, 12))

    # ── helpers ──────────────────────────────

    def _toggle_gs(self):
        if self.export_fmt.get() == "gsheet":
            self.gs_frame.pack(fill="x", padx=16, pady=2,
                               before=self.convert_btn.master)
        else:
            self.gs_frame.pack_forget()

    def _browse(self):
        path = filedialog.askopenfilename(
            filetypes=[("Images", "*.jpg *.jpeg *.png *.gif *.webp"),
                       ("All files", "*.*")])
        if path:
            self._load_image(path)

    def _browse_creds(self):
        path = filedialog.askopenfilename(
            filetypes=[("JSON", "*.json"), ("All files", "*.*")])
        if path:
            self.gs_creds.set(path)

    def _on_drop(self, event):
        path = event.data.strip().strip("{}")
        if path:
            self._load_image(path)

    def _load_image(self, path: str):
        self.image_path.set(path)
        self.status_msg.set(f"Loaded: {Path(path).name}")
        self.rows = []
        self._clear_tree()
        self.save_btn.config(state="disabled")

        if not HAS_PIL:
            return
        try:
            img = Image.open(path)
            img.thumbnail((860, 220), Image.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            self.preview_label.config(image=photo, text="", pady=0)
            self.preview_label.image = photo
        except Exception:
            pass

    def _start_conversion(self):
        path = self.image_path.get()
        if not path or not Path(path).exists():
            messagebox.showwarning("No image", "Please select an image file first.")
            return
        key = self.api_key.get().strip()
        if not key:
            messagebox.showwarning("API Key", "Please enter your Anthropic API key.")
            return

        self.convert_btn.config(state="disabled")
        self.save_btn.config(state="disabled")
        self.progress.start(10)
        self.status_msg.set("Sending image to Claude… this may take a moment.")
        threading.Thread(target=self._do_extraction,
                         args=(path, key), daemon=True).start()

    def _do_extraction(self, path: str, key: str):
        try:
            rows = extract_table_from_image(path, key)
            self.after(0, self._extraction_done, rows)
        except Exception as exc:
            self.after(0, self._extraction_error, str(exc))

    def _extraction_done(self, rows: list[list[str]]):
        self.progress.stop()
        self.convert_btn.config(state="normal")
        self.rows = rows
        self._populate_tree(rows)
        self.save_btn.config(state="normal")
        self.status_msg.set(
            f"Extracted {len(rows)} rows × {max(len(r) for r in rows)} columns. "
            "Ready to save.")

    def _extraction_error(self, msg: str):
        self.progress.stop()
        self.convert_btn.config(state="normal")
        self.status_msg.set(f"Error: {msg}")
        messagebox.showerror("Extraction failed", msg)

    def _clear_tree(self):
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = []

    def _populate_tree(self, rows: list[list[str]]):
        self._clear_tree()
        if not rows:
            return
        cols = [str(i) for i in range(len(rows[0]))]
        self.tree["columns"] = cols
        for c, heading in zip(cols, rows[0]):
            self.tree.heading(c, text=heading)
            self.tree.column(c, width=max(80, len(heading) * 9), anchor="center")
        for row in rows[1:]:
            # Pad short rows
            padded = row + [""] * (len(cols) - len(row))
            self.tree.insert("", "end", values=padded)

    def _save(self):
        if not self.rows:
            return
        fmt = self.export_fmt.get()
        if fmt == "csv":
            self._save_csv()
        elif fmt == "excel":
            self._save_excel()
        elif fmt == "gsheet":
            self._upload_gsheet()

    def _save_csv(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialfile="table.csv")
        if not path:
            return
        try:
            save_csv(self.rows, path)
            self.status_msg.set(f"Saved CSV → {path}")
            messagebox.showinfo("Saved", f"CSV saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Save failed", str(e))

    def _save_excel(self):
        if not HAS_OPENPYXL:
            messagebox.showerror("Missing package", "Install openpyxl: pip install openpyxl")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="table.xlsx")
        if not path:
            return
        try:
            save_excel(self.rows, path)
            self.status_msg.set(f"Saved Excel → {path}")
            messagebox.showinfo("Saved", f"Excel saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Save failed", str(e))

    def _upload_gsheet(self):
        if not HAS_GSPREAD:
            messagebox.showerror("Missing package",
                                 "Install gspread: pip install gspread google-auth")
            return
        creds = self.gs_creds.get().strip()
        if not creds or not Path(creds).exists():
            messagebox.showwarning("Credentials",
                                   "Please provide a valid service-account JSON file.\n\n"
                                   "See: console.cloud.google.com → Service Accounts")
            return
        title = self.gs_title.get().strip() or "Extracted Table"
        self.status_msg.set("Uploading to Google Sheets…")
        self.save_btn.config(state="disabled")
        threading.Thread(target=self._do_upload,
                         args=(title, creds), daemon=True).start()

    def _do_upload(self, title: str, creds: str):
        try:
            url = upload_google_sheets(self.rows, title, creds)
            self.after(0, self._upload_done, url)
        except Exception as e:
            self.after(0, self._upload_error, str(e))

    def _upload_done(self, url: str):
        self.save_btn.config(state="normal")
        self.status_msg.set(f"Uploaded → {url}")
        messagebox.showinfo("Google Sheets", f"Spreadsheet created:\n{url}")

    def _upload_error(self, msg: str):
        self.save_btn.config(state="normal")
        self.status_msg.set(f"Upload error: {msg}")
        messagebox.showerror("Upload failed", msg)


# ─────────────────────────────────────────────
#  Entry point
# ─────────────────────────────────────────────

if __name__ == "__main__":
    app = App()
    app.mainloop()
